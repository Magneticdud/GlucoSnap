import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from ..models import GlucoseReading, Meal


def export_to_csv(user):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="glucosnap_export.csv"'

    writer = csv.writer(response)

    # Readings
    writer.writerow(["--- Glucose Readings ---"])
    writer.writerow(["Date", "Time", "Level (mg/dL)", "Type", "Notes"])
    for reading in GlucoseReading.objects.filter(user=user):
        writer.writerow(
            [
                reading.timestamp.date(),
                reading.timestamp.time(),
                reading.glucose_level,
                reading.get_measurement_type_display(),
                reading.notes,
            ]
        )

    writer.writerow([])

    # Meals
    writer.writerow(["--- Meals ---"])
    writer.writerow(["Date", "Time", "Type", "Description", "Calories", "Carbs"])
    for meal in Meal.objects.filter(user=user):
        writer.writerow(
            [
                meal.timestamp.date(),
                meal.timestamp.time(),
                meal.get_meal_type_display(),
                meal.description or meal.manual_notes,
                meal.estimated_calories,
                meal.carbs_estimate,
            ]
        )

    return response


def export_to_excel(user):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="glucosnap_export.xlsx"'

    wb = openpyxl.Workbook()

    # Readings Sheet
    ws_readings = wb.active
    ws_readings.title = "Glucose Readings"

    headers = ["Date", "Time", "Level (mg/dL)", "Type", "Notes"]
    ws_readings.append(headers)

    # Style headers
    for cell in ws_readings[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )

    for reading in GlucoseReading.objects.filter(user=user):
        ws_readings.append(
            [
                reading.timestamp.date(),
                reading.timestamp.time(),
                reading.glucose_level,
                reading.get_measurement_type_display(),
                reading.notes,
            ]
        )

    # Meals Sheet
    ws_meals = wb.create_sheet("Meals")
    headers = ["Date", "Time", "Type", "Description", "Calories", "Carbs"]
    ws_meals.append(headers)

    for cell in ws_meals[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="E5FFCC", end_color="E5FFCC", fill_type="solid"
        )

    for meal in Meal.objects.filter(user=user):
        ws_meals.append(
            [
                meal.timestamp.date(),
                meal.timestamp.time(),
                meal.get_meal_type_display(),
                meal.description or meal.manual_notes,
                meal.estimated_calories,
                meal.carbs_estimate,
            ]
        )

    wb.save(response)
    return response


def export_to_ods(user):
    response = HttpResponse(
        content_type="application/vnd.oasis.opendocument.spreadsheet"
    )
    response["Content-Disposition"] = 'attachment; filename="glucosnap_export.ods"'

    doc = OpenDocumentSpreadsheet()

    # Readings Table
    table_readings = Table(name="Glucose Readings")

    # Headers
    tr = TableRow()
    for header in ["Date", "Time", "Level (mg/dL)", "Type", "Notes"]:
        tc = TableCell()
        tc.addElement(P(text=header))
        tr.addElement(tc)
    table_readings.addElement(tr)

    for reading in GlucoseReading.objects.filter(user=user):
        tr = TableRow()
        cells = [
            str(reading.timestamp.date()),
            str(reading.timestamp.time()),
            str(reading.glucose_level),
            reading.get_measurement_type_display(),
            reading.notes or "",
        ]
        for cell_val in cells:
            tc = TableCell()
            tc.addElement(P(text=cell_val))
            tr.addElement(tc)
        table_readings.addElement(tr)

    doc.spreadsheet.addElement(table_readings)

    # Meals Table
    table_meals = Table(name="Meals")

    tr = TableRow()
    for header in ["Date", "Time", "Type", "Description", "Calories", "Carbs"]:
        tc = TableCell()
        tc.addElement(P(text=header))
        tr.addElement(tc)
    table_meals.addElement(tr)

    for meal in Meal.objects.filter(user=user):
        tr = TableRow()
        cells = [
            str(meal.timestamp.date()),
            str(meal.timestamp.time()),
            meal.get_meal_type_display(),
            meal.description or meal.manual_notes or "",
            str(meal.estimated_calories or ""),
            str(meal.carbs_estimate or ""),
        ]
        for cell_val in cells:
            tc = TableCell()
            tc.addElement(P(text=cell_val))
            tr.addElement(tc)
        table_meals.addElement(tr)

    doc.spreadsheet.addElement(table_meals)

    doc.save(response)
    return response
